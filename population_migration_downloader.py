"""
행정안전부 주민등록 데이터 - 지역별 인구이동 현황 전체 다운로드
실행 방법: python population_migration_downloader.py
필요 패키지: pip install requests beautifulsoup4 openpyxl
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re
import time
import sys

# ─────────────────────────────────────────────────────────────
# 설정 (필요에 따라 수정)
# ─────────────────────────────────────────────────────────────
BASE_URL = "https://rdoa.jumin.go.kr/openStats/selectConPpltnData"

# 브라우저 개발자 도구의 쿠키 값을 아래에 붙여넣으세요 (로그인 세션 유지용)
# 예: "JSESSIONID=ABCDEF123456; 기타쿠키=값"
COOKIE_STRING = ""   # ← 필요 시 여기에 쿠키 붙여넣기

QUERY_PARAMS = {
    "paramUrl": "mvinAdmmCd=4119456000&mvtAdmmCd=1000000000&lv=3&srchFrYm=202512&srchToYm=202602",
    "lv": "3",
    "mvinCtpvCd": "4100000000",
    "mvinSggCd":  "4119000000",
    "mvinDongCd": "4119456000",
    "mvtCtpvCd":  "1000000000",
    "mvtSggCd":   "",
    "mvtDongCd":  "",
    "srchFrYear": "2025",
    "srchFrMon":  "12",
    "srchToYear": "2026",
    "srchToMon":  "02",
}

OUTPUT_FILE = "population_migration.xlsx"
DELAY_SEC   = 0.5   # 페이지 간 딜레이(초) — 서버 부하 방지
# ─────────────────────────────────────────────────────────────


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
        "content-type": "application/x-www-form-urlencoded",
        "referer": BASE_URL,
        "user-agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/146.0.0.0 Safari/537.36"
        ),
    })
    if COOKIE_STRING:
        for part in COOKIE_STRING.split(";"):
            part = part.strip()
            if "=" in part:
                k, v = part.split("=", 1)
                session.cookies.set(k.strip(), v.strip())
    return session


def fetch_page(page: int, session: requests.Session) -> str:
    body = {**QUERY_PARAMS, "curPage": str(page)}
    resp = session.post(BASE_URL, data=body, timeout=30)
    resp.raise_for_status()
    resp.encoding = "utf-8"
    return resp.text


def parse_page(html: str):
    soup = BeautifulSoup(html, "html.parser")

    # 총 건수
    total_elem = soup.select_one("p.total b")
    total_count = int(total_elem.text.replace(",", "")) if total_elem else 0

    # 마지막 페이지 번호
    last_page = 1
    last_link = soup.select_one("a.last")
    if last_link:
        m = re.search(r"clickPage\((\d+)\)", last_link.get("onclick", ""))
        if m:
            last_page = int(m.group(1))

    # 헤더
    headers = [th.get_text(strip=True) for th in soup.select("table thead tr th")]

    # 데이터 행
    rows = []
    for tr in soup.select("table tbody tr"):
        row = [td.get_text(strip=True) for td in tr.find_all("td")]
        if row:
            rows.append(row)

    return headers, rows, total_count, last_page


def save_excel(headers: list, all_rows: list, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "인구이동현황"

    # ── 헤더 스타일
    h_font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    h_fill  = PatternFill("solid", start_color="2F5496")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font  = h_font
        cell.fill  = h_fill
        cell.alignment = h_align

    ws.row_dimensions[1].height = 35

    # ── 데이터
    d_font  = Font(name="Arial", size=9)
    d_align = Alignment(horizontal="center", vertical="center")

    for r_idx, row_data in enumerate(all_rows, 2):
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font      = d_font
            cell.alignment = d_align

    # ── 열 너비 (주요 9개 컬럼)
    key_widths = [10, 16, 16, 12, 16, 12, 12, 16, 14]
    for i, w in enumerate(key_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(10, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 5

    # ── 헤더 고정
    ws.freeze_panes = "A2"

    wb.save(path)


def main():
    session = build_session()

    print("[ 1/? ] 1페이지 수집 중...")
    try:
        html = fetch_page(1, session)
    except requests.RequestException as e:
        print(f"[오류] 요청 실패: {e}")
        sys.exit(1)

    headers, rows, total_count, last_page = parse_page(html)

    if not headers:
        print("[오류] 헤더를 파싱하지 못했습니다. 쿠키 또는 URL을 확인하세요.")
        sys.exit(1)

    print(f"       총 {total_count:,}건 / {last_page}페이지")
    all_rows = list(rows)

    for page in range(2, last_page + 1):
        print(f"[ {page}/{last_page} ] 수집 중...", end="\r", flush=True)
        try:
            html = fetch_page(page, session)
        except requests.RequestException as e:
            print(f"\n[경고] {page}페이지 실패 (건너뜀): {e}")
            continue
        _, rows, _, _ = parse_page(html)
        all_rows.extend(rows)
        time.sleep(DELAY_SEC)

    print(f"\n수집 완료: {len(all_rows):,}행")

    print(f"엑셀 저장 중: {OUTPUT_FILE} ...")
    save_excel(headers, all_rows, OUTPUT_FILE)
    print(f"완료! → {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
