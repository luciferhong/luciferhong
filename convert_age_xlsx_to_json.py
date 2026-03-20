from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


def normalize_header_name(value: Any) -> str:
    return str(value or "").replace(" ", "").replace("~", "-").strip()


def to_number(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "")
    text = text.replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except Exception:
        return 0.0


def resolve_age_band_indexes(headers: list[str]) -> dict[str, list[int]]:
    out = {
        "band0": [],
        "band10": [],
        "band20": [],
        "band30": [],
        "band50": [],
        "band60": [],
    }

    def starts_with_age(h: str, start: int, end: int) -> bool:
        return h.startswith(f"{start}-{end}") or h.startswith(f"{start}세{end}세")

    for idx, raw in enumerate(headers):
        h = normalize_header_name(raw)
        if not h:
            continue
        if starts_with_age(h, 0, 9) or "9세이하" in h:
            out["band0"].append(idx)
        if starts_with_age(h, 10, 19):
            out["band10"].append(idx)
        if starts_with_age(h, 20, 29):
            out["band20"].append(idx)
        if h.startswith("30-49") or h.startswith("30세49세") or starts_with_age(h, 30, 39) or starts_with_age(h, 40, 49):
            out["band30"].append(idx)
        if starts_with_age(h, 50, 59):
            out["band50"].append(idx)
        if (
            "60세이상" in h
            or starts_with_age(h, 60, 69)
            or starts_with_age(h, 70, 79)
            or starts_with_age(h, 80, 89)
            or starts_with_age(h, 90, 99)
            or h.startswith("100세이상")
            or h.startswith("100-")
        ):
            out["band60"].append(idx)

    return out


def pick_label_index(headers: list[str]) -> int:
    preferred = ["법정동", "읍면동", "행정동", "동", "지역", "행정기관"]
    for key in preferred:
        for idx, h in enumerate(headers):
            if key in str(h):
                return idx
    return 0


def load_emd_code_name_map(base: Path) -> dict[str, str]:
    mapping_path = base / "emd_region_codes.xlsx"
    if not mapping_path.exists():
        return {}

    wb = openpyxl.load_workbook(mapping_path, data_only=True, read_only=True)
    ws = wb.active

    out: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 2:
            continue
        code = str(row[0] or "").strip()
        name = str(row[1] or "").strip()
        if code and name:
            out[code] = name
    return out


def main() -> None:
    base = Path(__file__).resolve().parent
    code_name_map = load_emd_code_name_map(base)
    src = base / "age_population_emd.xlsx"
    if not src.exists():
        src = base / "법정동연령별인구.xlsx"
    if not src.exists():
        raise FileNotFoundError("No source workbook found. Expected age_population_emd.xlsx or 법정동연령별인구.xlsx")

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Workbook is empty")

    headers = [str(v or "").strip() for v in rows[0]]
    label_idx = pick_label_index(headers)
    band_idx = resolve_age_band_indexes(headers)

    out_rows: list[dict[str, Any]] = []
    for row in rows[1:]:
        raw_label = str(row[label_idx] or "").strip() if label_idx < len(row) else ""
        if not raw_label:
            continue

        # 라벨이 코드(숫자)인 경우 동명으로 치환
        label_code = "".join(ch for ch in raw_label if ch.isdigit())
        code8 = label_code[:8] if len(label_code) >= 8 else label_code
        label = code_name_map.get(label_code) or code_name_map.get(code8) or raw_label

        def sum_by_indexes(indexes: list[int]) -> float:
            total = 0.0
            for i in indexes:
                if i < len(row):
                    total += to_number(row[i])
            return total

        v0 = sum_by_indexes(band_idx["band0"])
        v10 = sum_by_indexes(band_idx["band10"])
        v20 = sum_by_indexes(band_idx["band20"])
        v30 = sum_by_indexes(band_idx["band30"])
        v50 = sum_by_indexes(band_idx["band50"])
        v60 = sum_by_indexes(band_idx["band60"])

        total = v0 + v10 + v20 + v30 + v50 + v60
        if total <= 0:
            continue

        out_rows.append(
            {
                "label": label,
                "code": label_code,
                "values": [v0, v10, v20, v30, v50, v60],
                "total": total,
            }
        )

    payload = {
        "source": src.name,
        "ageBands": ["0~9", "10~19", "20~29", "30~49", "50~59", "60+"],
        "rows": out_rows,
    }

    dst = base / "age_population_emd.json"
    dst.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    print(f"Wrote {dst} with {len(out_rows)} rows")


if __name__ == "__main__":
    main()
